from openpyxl import load_workbook, Workbook

# helper
def get_headers(filePath):
    ' Get excel headers '
    sheet = load_workbook(filePath).worksheets[0]
    headers = [sheet.cell(sheet.min_row, i).value for i in range(sheet.min_column, sheet.max_column+1)]
    return headers

def do_sampling(sourceFile, config):
    ' Sample sourcefile according to config '
    ' config = {"size":3, "useSize":True, "rate":4, "useRate":False} '
    sheet = load_workbook(sourceFile).worksheets[0]
    # sampling
    # dpts = sheet.col_values(int(config['dropDown']))
    dpt_col = int(config['dropDown'])+sheet.min_column
    dpts = [sheet.cell(i, dpt_col).value for i in range(sheet.min_row, sheet.max_row+1)]
    indices = []
    sampleSize = config['size'] if config['useSize'] else 0
    sampleRate = config['rate'] if config['useRate'] else 0

    nEntry = sheet.max_row - sheet.min_row + 1
    start = 1
    while start < nEntry:
        end = start
        while end < nEntry and dpts[start] == dpts[end]:
            end += 1

        size = end - start
        realSampleSize = max(sampleSize, sampleRate*size/100)
        step = max(size / realSampleSize, 1)
        for i in range(int(realSampleSize)):
            val = int(start + i * step)
            if val >= end:
                break
            indices.append(val)
        start = end
    
    # create workbook
    result = Workbook()
    targetSheet = result.active
    # write
    targetSheet.cell(1, 1).value = '序号'
    ncols = sheet.max_column - sheet.min_column + 1
    for i in range(ncols):
        targetSheet.cell(1, i+2).value = sheet.cell(sheet.min_row, i+sheet.min_column).value

    targetRow = 2
    for i in indices:
        targetSheet.cell(targetRow, 1).value = targetRow - 1
        for j in range(ncols):
            sourceCell = sheet.cell(i+sheet.min_row, j+sheet.min_column)
            targetSheet.cell(targetRow, j+2).value = sourceCell.value
            targetSheet.cell(targetRow, j+2).data_type = sourceCell.data_type
        targetRow += 1
    return result

